import openpyxl as xl
import json
import pprint
import sys
import numpy as np

#user Input
# file_name = input("what excel file you want to use?")
sheet_name = input("which sheet do you want to convert?")
output_file = input("what is the name of json file (please enter the name with .json) ?")


##########################################################
################# Variables ##############################
##########################################################
#define the variable
column_number = 1
count_col = 0
count_row = 0
row_number = 1

output_list = []
key = []
merge_list=[]

##########################################################
################# Functions ##############################
##########################################################
#Build the dirctionary of unmerged file
def unmerged():
    for r in range(2,(count_row+1)):
        output_list.append({})
        for c in range(1,column_number):
            output_list[(r-2)].update({key[(c-1)]:ws.cell(row = r,column = c).value})
    return output_list

def check_merge():
    merge = ws.merged_cells.ranges
    if not merge:
        print("\n Congraduation! No merged cell\n")
    else:
        for list in merge:
            merge_list.append(str(list))
        return merge_list
#merge_list should be ['A2:A3','B2:B3']
def merged():
    cause_merge_number = len(merge_list)/2   #number of how many causes has merged cells
    valid_row = 0
    count = -1
    temp = []
    for r in range(2,(count_row+1)):
        if(ws.cell(row = r,column = 1).value != None):
            output_list.append({})
            output_list[valid_row].update({key[0]:ws.cell(row = r,column = 1).value})
            output_list[valid_row].update({key[1]:ws.cell(row = r,column = 2).value})
            valid_row = valid_row + 1
    for r in range(2,(count_row+1)):
        if(ws.cell(row = r,column = 1).value != None):
            count = count + 1
        for c in range(3,column_number):
            if(ws.cell(row = r,column = 1).value == None):
                value = output_list[count][key[(c-1)]]
                if(isinstance(value,(list,tuple))):
                    print("here we go")
                    temp = value
                else:
                    temp.append(value)
                temp.append(ws.cell(row = r,column = c).value)
                output_list[count][key[(c-1)]] = temp
                temp =[]
            else:
                output_list[count].update({key[(c-1)]:ws.cell(row = r,column = c).value})

    return output_list
##########################################################
################# Execute ##############################
##########################################################
#load the excel file
#wb = xl.load_workbook(file_name,data_only = True)
wb = xl.load_workbook("Name.xlsx",data_only = True)
#get the sheet
ws = wb[sheet_name]
# ws = wb["Cooling System"]
check_merge()
#read the title of each row.
for col in ws.columns:
    # Fields.append([])
    # Fields[count_col].append(ws.cell(row=1,column=column_number).value) #17
    key.append(ws.cell(row=1,column=column_number).value)
    column_number = column_number+1    #18
    count_col = count_col +1 #17
print("titles: ",key,'\n')
print(column_number)
#count row
count_row = ws.max_row

if not merge_list:
    unmerged()
else:
    merged()
#check whether there is a blank row
remove_list = []
for l in range(len(output_list)):
    if(output_list[l][key[0]] == None):
        remove_list.append(l)

output_list = np.delete(output_list,remove_list).tolist()

print("# Valid Row:",len(output_list),'\n')
print("# of Column:",len(output_list[1]),'\n')
#print(output_list)


with open(output_file,'w') as f:
    #json.dump(jsonData, f)
    f.write(json.dumps(output_list,indent=4,sort_keys=True))
